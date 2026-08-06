#!/usr/bin/env python3
"""Extract protocol-preserving Wave F evidence for Hibou, MUSK, and GPFM.

This evidence-only extractor does not rebuild the shared registry.  It pins the
official papers/workbooks, keeps checkpoint and protocol families separate, and
quarantines aggregates, task-fine-tuned systems, and non-public cohorts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import subprocess
from pathlib import Path

FIELDS = (
    "source_scope", "source_table", "source_row", "model_alias", "model_id",
    "model_revision", "suite_id", "task_label", "evaluation_id", "dedup_key",
    "metric", "value", "value_unit", "uncertainty", "level", "magnification",
    "embedding_recipe", "downstream_protocol", "cohort_access", "reference_url",
    "source_locator", "source_revision", "source_sha256", "inclusion_status",
    "inclusion_reason",
)

HIBOU_SHA = "5c4086cad4dfa47ae6699a53149362a9a593785830b56dcc4971a72fe95d5fe5"
MUSK_SHA = "f96a44a9e3b531472a166a0a06dc3dec241b3f75fc455a21479c798f6028b770"
GPFM_SUPP_SHA = "7fb834ee12f33fcdd369fe8f218c82a13523f12401f3bac8d24c10bee2f77b2f"
GPFM_XLSX_SHA = {
    4: "0431fdab9f8c5f410fb0c5ae77103a86158ecce961d6631a291f8f086a2bfb97",
    5: "145f5a913a36d79a8fd13b9b4783817e5cad3dcdbeecf4fc69857cdceff184c4",
    6: "97742abaaefebc70435d04948eaa9d4d4d6b9447b982baf26d587fb85d5f0d76",
    7: "544dbdc7a0e2cccf0642da377af682b3b979aee765e88d78e6c13e1d3965fb2d",
    8: "0f62bd4e74a8736286a82d562d122603bce49ea2b37daac2f84bcb4706f2dfd9",
    9: "71aa5171037b97c36d8b9a1f1222c11ed6a0f79671393ac6265a218a207c76c9",
    10: "10b4f275bfb98762c883b89b173c87d187e6a32c75f35352b19f44b3d9530b71",
    11: "acabd754d8a63594314bce1ae3281041fd5a2401da602929a6723a7886f93122",
    12: "71777ead89466460982c295109dfd3ecf0e7a9770265b2259ffc715b02387a67",
    13: "5d33aa355859a25b7f133865904851e87aa58beb4afcdb259dc0bb36e56996f3",
    14: "56a270d8e4dcce21a63e98d540b47e3a0b97c68cca7e838595138cca75ab7fa1",
}

HIBOU_URL = "https://arxiv.org/abs/2406.05074"
MUSK_URL = "https://www.nature.com/articles/s41586-024-08378-w"
GPFM_URL = "https://www.nature.com/articles/s41551-025-01488-4"

# The supplement states that zero-shot classification covers six datasets, but
# Supplementary Table 4 publishes exact leaves for only four.  NCT-CRC and
# SICAPv2 are named as zero-shot datasets in the methods/main-paper experiment,
# but no exact first-party numeric leaf is present in the supplement or repo.
MUSK_ZERO_SHOT_DISPOSITION = {
    "PatchCamelyon": "exact_table4_leaf",
    "SkinCancer": "exact_table4_leaf",
    "PanNuke": "exact_table4_leaf",
    "UniToPatho": "exact_table4_leaf",
    "NCT-CRC": "graph_only_unlocated_exact_value",
    "SICAPv2": "graph_only_unlocated_exact_value",
}


def _digest(path: Path, expected: str) -> str:
    actual = hashlib.sha256(path.read_bytes()).hexdigest()
    if actual != expected:
        raise ValueError(f"unexpected SHA-256 for {path}: {actual}")
    return actual


def _text(path: Path, expected: str) -> str:
    _digest(path, expected)
    return subprocess.run(
        ["pdftotext", "-layout", str(path), "-"], check=True, text=True,
        stdout=subprocess.PIPE, stderr=subprocess.PIPE,
    ).stdout


def _row(**kwargs: str) -> dict[str, str]:
    missing = set(FIELDS) - set(kwargs)
    if missing:
        raise ValueError(f"missing fields: {sorted(missing)}")
    return {key: kwargs[key] for key in FIELDS}


def _write(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def _paper_blocks(text: str) -> dict[int, str]:
    matches = list(re.finditer(r"Supplementary Table (\d+):", text))
    return {
        int(match.group(1)): text[match.start():(matches[i + 1].start() if i + 1 < len(matches) else len(text))]
        for i, match in enumerate(matches)
    }


def _common(model_alias: str, model_id: str, revision: str, suite: str, url: str,
            source_revision: str, sha: str) -> dict[str, str]:
    return {
        "model_alias": model_alias, "model_id": model_id, "model_revision": revision,
        "suite_id": suite, "reference_url": url, "source_revision": source_revision,
        "source_sha256": sha,
    }


def hibou_rows(pdf: Path) -> list[dict[str, str]]:
    text = _text(pdf, HIBOU_SHA)
    common = _common("", "", "", "hibou_primary", HIBOU_URL,
                     "arXiv:2406.05074v1 (2024-06-07)", HIBOU_SHA)
    rows: list[dict[str, str]] = []
    patch = {
        "CRC-100K": ("0.955", "0.966"), "PCAM": ("0.946", "0.953"),
        "MHIST": ("0.812", "0.858"), "MSI-CRC": ("0.779", "0.793"),
        "MSI-STAD": ("0.797", "0.829"), "TIL-DET": ("0.942", "0.942"),
    }
    slide = {"BRCA": ("0.929", "0.946"), "NSCLC": ("0.952", "0.969"), "RCC": ("0.993", "0.996")}
    for model, index in (("hibou-b", 0), ("hibou-l", 1)):
        alias = "Hibou-B" if model == "hibou-b" else "Hibou-L"
        revision = f"{alias} ViT-{'B' if model == 'hibou-b' else 'L'}/14; official arXiv v1 checkpoint family"
        for task, values in patch.items():
            value = values[index]
            if f"{task}" not in text or value not in text:
                raise ValueError(f"Hibou Table 1 source changed for {model}/{task}")
            rows.append(_row(
                source_scope="official_primary_paper", source_table="Table 1",
                source_row=f"{alias}/{task}/top-1 accuracy", model_alias=alias,
                model_id=model, model_revision=revision, suite_id=common["suite_id"],
                task_label=task, evaluation_id=f"{model}.hibou2024.t1.{_slug(task)}.top1_accuracy",
                dedup_key=f"{_slug(task)}.linear_probe.top1_accuracy", metric="top1_accuracy",
                value=value, value_unit="fraction", uncertainty="", level="tile",
                magnification="source-defined; 224 px input (PCam upsampled)",
                embedding_recipe=f"frozen {alias} CLS embedding",
                downstream_protocol="supervised linear probe; SGD; validation-selected linear head",
                cohort_access="public", reference_url=common["reference_url"],
                source_locator=f"paper|Table 1|{alias}|{task}", source_revision=common["source_revision"],
                source_sha256=common["source_sha256"], inclusion_status="canonical_candidate",
                inclusion_reason="Public leaf result with a frozen foundation encoder and linear head.",
            ))
        for task, values in slide.items():
            value = values[index]
            rows.append(_row(
                source_scope="official_primary_paper", source_table="Table 2",
                source_row=f"{alias}/{task}/AUC", model_alias=alias, model_id=model,
                model_revision=revision, suite_id=common["suite_id"], task_label=task,
                evaluation_id=f"{model}.hibou2024.t2.{_slug(task)}.auroc",
                dedup_key=f"tcga_{_slug(task)}.frozen_abmil.auroc", metric="auroc", value=value,
                value_unit="fraction", uncertainty="", level="slide", magnification="source-defined non-overlapping foreground patches",
                embedding_recipe=f"frozen {alias} patch embeddings; learned attention pooling",
                downstream_protocol="weakly supervised attention pooling; encoder frozen; 80/10/10 split",
                cohort_access="public", reference_url=common["reference_url"],
                source_locator=f"paper|Table 2|{alias}|{task}", source_revision=common["source_revision"],
                source_sha256=common["source_sha256"], inclusion_status="canonical_candidate",
                inclusion_reason="Public slide-level leaf with frozen patch encoder.",
            ))

    pq = {"neoplastic": "0.582", "epithelial": "0.591", "inflammatory": "0.426", "connective": "0.425", "dead": "0.185"}
    prf = {
        "neoplastic": ("0.72", "0.72", "0.72"), "epithelial": ("0.76", "0.76", "0.76"),
        "inflammatory": ("0.63", "0.58", "0.60"), "connective": ("0.58", "0.53", "0.55"),
        "dead": ("0.51", "0.35", "0.41"),
    }
    for category, value in pq.items():
        prf_values = (("panoptic_quality", value), *zip(("precision", "recall", "f1"), prf[category]))
        for metric, cell in prf_values:
            table = "Table 3" if metric == "panoptic_quality" else "Table 4"
            rows.append(_row(
                source_scope="official_primary_paper", source_table=table,
                source_row=f"CellViT-Hibou-L/{category}/{metric}", model_alias="CellViT-Hibou-L",
                model_id="hibou-l", model_revision="Hibou-L backbone embedded in task-trained CellViT",
                suite_id=common["suite_id"], task_label=f"PanNuke nuclei/{category}",
                evaluation_id=f"hibou-l.hibou2024.{table.lower().replace(' ', '')}.pannuke.{category}.{metric}",
                dedup_key=f"pannuke.{category}.{metric}", metric=metric, value=cell, value_unit="fraction",
                uncertainty="mean across three PanNuke splits", level="tile", magnification="source-defined",
                embedding_recipe="Hibou-L used as CellViT backbone",
                downstream_protocol="end-to-end nuclei instance segmentation/classification training",
                cohort_access="public", reference_url=common["reference_url"],
                source_locator=f"paper|{table}|CellViT-Hibou-L|{category}", source_revision=common["source_revision"],
                source_sha256=common["source_sha256"], inclusion_status="fine_tuned_excluded",
                inclusion_reason="Task-specific CellViT training changes the backbone and is not a frozen-encoder probe.",
            ))
    if len(rows) != 38:
        raise ValueError(f"expected 38 Hibou leaves, got {len(rows)}")
    return rows


CI_PERCENT = re.compile(r"(\d+(?:\.\d+)?)\s*\(\s*(\d+(?:\.\d+)?)\s*,\s*(\d+(?:\.\d+)?)\s*\)")
SD_PERCENT = re.compile(r"(\d+(?:\.\d+)?)\s*\(\s*(\d+(?:\.\d+)?)\s*\)")
CI_FRACTION = re.compile(r"(0?\.\d+|1(?:\.0+)?)\s*\(\s*(0?\.\d+|1(?:\.0+)?)\s*,\s*(0?\.\d+|1(?:\.0+)?)\s*\)")


def musk_rows(pdf: Path) -> list[dict[str, str]]:
    text = _text(pdf, MUSK_SHA)
    blocks = _paper_blocks(text)
    common = _common("MUSK", "musk", "MUSK-large ViT-L/16, 675M parameters",
                     "musk_primary", MUSK_URL, "Nature 638 (2025) supplementary PDF", MUSK_SHA)
    rows: list[dict[str, str]] = []

    def add(table: int, task: str, metric: str, value: str, uncertainty: str, protocol: str,
            level: str = "tile", status: str = "canonical_candidate", access: str = "public",
            reason: str = "Public primary-paper leaf protocol cell.", alias: str = "MUSK",
            component_id: str = "musk") -> None:
        rows.append(_row(
            source_scope="official_primary_paper", source_table=f"Supplementary Table {table}",
            source_row=f"{alias}/{task}/{metric}", model_alias=alias, model_id=component_id,
            model_revision=common["model_revision"] if alias == "MUSK" else alias,
            suite_id=common["suite_id"], task_label=task,
            evaluation_id=f"musk.nature2025.t{table}.{_slug(alias)}.{_slug(task)}.{metric}",
            dedup_key=f"{_slug(task)}.{_slug(protocol)}.{metric}", metric=metric, value=value,
            value_unit="fraction" if table == 7 else "percent", uncertainty=uncertainty,
            level=level, magnification="source-defined; 384x384 MUSK input",
            embedding_recipe="MUSK image/text embeddings" if level == "tile" else "MUSK multimodal patient representation",
            downstream_protocol=protocol, cohort_access=access, reference_url=common["reference_url"],
            source_locator=f"Supplementary Information|Table {table}|{alias}|{task}",
            source_revision=common["source_revision"], source_sha256=common["source_sha256"],
            inclusion_status=status, inclusion_reason=reason,
        ))

    for table, protocol in ((1, "zero-shot text-to-image retrieval"), (2, "zero-shot image-to-text retrieval")):
        lines = [line for line in blocks[table].splitlines() if re.search(r"\bMUSK\b", line)]
        if len(lines) != 2:
            raise ValueError(f"MUSK table {table} inventory changed")
        for task, line in zip(("BookSet", "PathMMU"), lines, strict=True):
            cells = CI_PERCENT.findall(line)
            for metric, (value, low, high) in zip(("recall_at_1", "recall_at_10", "recall_at_50"), cells, strict=True):
                add(table, task, metric, value, f"95% CI [{low}, {high}]", protocol)
    lines = [line for line in blocks[3].splitlines() if re.search(r"\bMUSK\b", line)]
    for task, line in zip(("BRACS-6", "UniToPatho"), lines, strict=True):
        for metric, (value, low, high) in zip(("recall_at_1", "recall_at_3", "recall_at_5", "majority_vote_at_5"), CI_PERCENT.findall(line), strict=True):
            add(3, task, metric, value, f"95% CI [{low}, {high}]", "zero-shot image-to-image retrieval")
    line = next(line for line in blocks[4].splitlines() if re.search(r"\bMUSK\b", line))
    for task, (value, low, high) in zip(("PatchCamelyon", "SkinCancer", "PanNuke", "UniToPatho"), CI_PERCENT.findall(line), strict=True):
        add(4, task, "balanced_accuracy", value, f"95% CI [{low}, {high}]", "zero-shot image classification; text prompts")
    tasks12 = ("NCT-CRC", "PatchCamelyon", "SkinCancer", "SICAPv2", "PanNuke", "UniToPatho", "WSSS4LUAD", "Osteo", "LC25000", "RenalCell", "BRACS-6", "BRACS-3")
    for table, protocol in ((5, "10-shot per class supervised classification; frozen encoder"), (6, "supervised linear probe; frozen encoder")):
        for task in tasks12:
            source_name = {"BRACS-6": "BRCAS (6-class)", "BRACS-3": "BRCAS (3-class)"}.get(task, task)
            line = next(line for line in blocks[table].splitlines() if re.match(rf"\s*{re.escape(source_name)}", line, re.I))
            value, sd = SD_PERCENT.findall(line)[-1]
            add(table, task, "balanced_accuracy", value, f"reported SD {sd}; 10 independent experiments", protocol)

    table7 = blocks[7]
    labels = ("BCNB-ER", "BCNB-PR", "BCNB-HER2", "MUV-IDH", "BLCA", "BRCA", "CESC", "COADREAD", "ESCA", "GBM", "HNSC", "LGG", "LIHC", "LUAD", "LUSC", "PAAD", "RCC", "SKCM", "STAD", "UCEC", "Lung-AUC", "Lung-c-index", "Gastro-esophageal-AUC", "Gastro-esophageal-c-index")
    for label in labels:
        pattern = label.replace("Gastro-esophageal", "Gastro-eso.").replace("Lung-AUC", "Lung (AUC)").replace("Lung-c-index", "Lung (c-index)").replace("-AUC", "(AUC)").replace("-c-index", "(c-index)")
        line = next(line for line in table7.splitlines() if pattern in line)
        value, low, high = CI_FRACTION.findall(line)[-1]
        private = label.startswith(("Lung-", "Gastro-"))
        metric = "auroc" if label.endswith("AUC") or label.startswith(("BCNB", "MUV")) else "c_index"
        protocol = "five-fold multimodal outcome model" if private else ("five-fold ABMIL biomarker prediction" if label.startswith(("BCNB", "MUV")) else "five-fold multimodal prognosis model")
        add(7, label, metric, value, f"95% CI [{low}, {high}]", protocol, level="slide",
            status="private_internal_excluded" if private else "canonical_candidate",
            access="controlled" if private else "public",
            reason="Controlled-access immunotherapy cohort." if private else "Public leaf slide-level outcome result.")

    aggregate_tasks = ("linear probe classification", "10-shot classification", "zero-shot classification", "image-to-image retrieval", "text-to-image retrieval", "image-to-text retrieval", "VQA")
    for table, count, aliases in ((13, 1, ("MUSK-ablation-final",)), (14, 4, tuple(f"MUSK-ablation-{i}" for i in range(1, 5))), (15, 1, ("MUSK-base-model5",)), (17, 3, ("MUSK-small-75M", "MUSK-base-222M", "MUSK-large-675M"))):
        block = blocks[table]
        task_lines = [line for line in block.splitlines() if re.match(
            r"\s*(?:Ablation \d+\s+)?(?:Linear probe cls|10-shot cls|Zero-shot cls|I2I retrieval|T2I retrieval|I2T retrieval|VQA)\s+\d", line
        )]
        expected_lines = 28 if table == 14 else 7
        if len(task_lines) != expected_lines:
            raise ValueError(f"MUSK table {table} aggregate inventory changed: {len(task_lines)}")
        groups = [task_lines[i * 7:(i + 1) * 7] for i in range(count)] if table == 14 else [task_lines]
        for group_index, group in enumerate(groups):
            for task, line in zip(aggregate_tasks, group, strict=True):
                cells = re.findall(r"\d+(?:\.\d+)?(?:\(\d+(?:\.\d+)?\))?", line)
                values = [cell for cell in cells if "." in cell]
                chosen = values[-2] if table in {13, 14} else values[-1]
                targets = aliases if table == 17 else (aliases[group_index],)
                for alias_index, alias in enumerate(targets):
                    cell = values[alias_index] if table == 17 else chosen
                    match = re.match(r"([\d.]+)(?:\(([\d.]+)\))?", cell)
                    assert match
                    component_ids = {
                        "MUSK-ablation-final": "musk-ablation-final",
                        "MUSK-ablation-1": "musk-ablation-1", "MUSK-ablation-2": "musk-ablation-2",
                        "MUSK-ablation-3": "musk-ablation-3", "MUSK-ablation-4": "musk-ablation-4",
                        "MUSK-base-model5": "musk-base-model5", "MUSK-small-75M": "musk-small",
                        "MUSK-base-222M": "musk-base", "MUSK-large-675M": "musk-large",
                    }
                    is_vqa = task == "VQA"
                    add(table, "PathVQA" if is_vqa else task, "accuracy" if is_vqa else "reported_task_metric",
                        match.group(1), f"reported aggregate SD {match.group(2)}" if match.group(2) else "",
                        "whole-model and classification-head fine-tuning on PathVQA" if is_vqa else "aggregate across datasets / ablation setting",
                        status="fine_tuned_excluded" if is_vqa else "aggregate_excluded", access="public",
                        reason="The methods explicitly fine-tune the complete MUSK backbone and answer head on PathVQA."
                        if is_vqa else "Aggregate or ablation cell is not an independent benchmark leaf.",
                        alias=alias, component_id=component_ids[alias])
    return rows


GPFM_PUBLIC_WSI = {
    "TCGA-NSCLC", "CPTAC-NSCLC", "TCGA-RCC", "CAMELYON", "TCGA-BRCA_IDC_ILC",
    "BRACS-3", "BRCAS7", "PANDA", "TP53", "IDH1_TCGA", "UBC-OCEAN",
    "TCGA-GBMLGG", "IMP", "Lauren-TCGA-STAD",
}
GPFM_PUBLIC_ROI = {
    "CRC100", "CCRCC-TCGA-HEL", "bach", "breakhis", "UniToPatho", "CRC-MSI",
    "PanCancer-TCGA", "PanCancer-TIL", "ESCA", "PCAM", "WSSS4LUAD", "Chaoyang", "GasHisDB",
}


def _gpfm_ci_groups(text: str, tables: range) -> list[tuple[str, str, str]]:
    blocks = _paper_blocks(text)
    out: list[tuple[str, str, str]] = []
    pattern = re.compile(r"(0?\.\d+|1(?:\.0+)?)\s*\(\s*(0?\.\d+|1(?:\.0+)?)\s*[-–]\s*(0?\.\d+|1(?:\.0+)?)\s*\)")
    for table in tables:
        for line in blocks[table].splitlines():
            if re.search(r"\bGPFM\b", line):
                out.extend(pattern.findall(line))
    return out


def gpfm_rows(supplement: Path, workbook_dir: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "GPFM workbook extraction requires openpyxl; install it before "
            "rebuilding the pinned evidence snapshot"
        ) from exc

    text = _text(supplement, GPFM_SUPP_SHA)
    paths = {n: workbook_dir / f"41551_2025_1488_MOESM{n}_ESM.xlsx" for n in GPFM_XLSX_SHA}
    for n, path in paths.items():
        _digest(path, GPFM_XLSX_SHA[n])
    common = _common("GPFM", "gpfm", "GPFM ViT-L/14 UDK checkpoint; publisher model family",
                     "gpfm_primary", GPFM_URL, "Nature Biomedical Engineering version of record 2025-09-02", GPFM_SUPP_SHA)
    rows: list[dict[str, str]] = []

    def add(source_table: str, source_row: str, task: str, metric: str, value: object,
            uncertainty: str, level: str, protocol: str, status: str, access: str,
            reason: str, workbook: int, sheet: str) -> None:
        rows.append(_row(
            source_scope="official_publisher_source_data", source_table=source_table,
            source_row=source_row, model_alias=common["model_alias"], model_id=common["model_id"],
            model_revision=common["model_revision"], suite_id=common["suite_id"], task_label=task,
            evaluation_id=f"gpfm.nbe2025.{_slug(source_table)}.{_slug(task)}.{metric}",
            dedup_key=f"{_slug(task)}.{_slug(protocol)}.{metric}", metric=metric,
            value=str(value), value_unit="percent" if "VQA" in source_table and float(value) > 1 else "fraction",
            uncertainty=uncertainty, level=level, magnification="source-defined",
            embedding_recipe="frozen GPFM patch embeddings" if "fine-tun" not in protocol else "task-fine-tuned GPFM system",
            downstream_protocol=protocol, cohort_access=access, reference_url=common["reference_url"],
            source_locator=f"Source Data workbook MOESM{workbook}|sheet {sheet}|{source_row}",
            source_revision=common["source_revision"], source_sha256=GPFM_XLSX_SHA[workbook],
            inclusion_status=status, inclusion_reason=reason,
        ))

    # Published means are controlled by the source workbooks; confidence intervals
    # are aligned in publisher-table row order from the supplement.
    wsi_ci = _gpfm_ci_groups(text, range(2, 19))
    wb = load_workbook(paths[6], read_only=True, data_only=True)
    wsi_rows = list(wb["b"].iter_rows(min_row=2, values_only=True))
    if len(wsi_rows) != 36 or len(wsi_ci) != 108:
        raise ValueError(f"GPFM WSI inventory changed: {len(wsi_rows)} rows/{len(wsi_ci)} cells")
    for row_index, row in enumerate(wsi_rows):
        task = str(row[0])
        public = task in GPFM_PUBLIC_WSI
        for metric_index, (sheet, metric) in enumerate((("b", "balanced_accuracy"), ("c", "weighted_f1"), ("d", "auroc"))):
            value = load_workbook(paths[6], read_only=True, data_only=True)[sheet].cell(row_index + 2, 10).value
            _, low, high = wsi_ci[row_index * 3 + metric_index]
            add("WSI classification", f"{task}/{metric}", task, metric, value, f"95% CI [{low}, {high}]", "slide",
                "two-layer ABMIL on frozen GPFM embeddings", "canonical_candidate" if public else "private_internal_excluded",
                "public" if public else "private_internal_or_unlisted",
                "Public frozen-encoder WSI leaf." if public else "Cohort is private/internal or absent from the paper's public-data table.", 6, sheet)

    survival_ci = _gpfm_ci_groups(text, range(20, 24))
    wb7 = load_workbook(paths[7], read_only=True, data_only=True)
    surv = list(wb7["b"].iter_rows(min_row=2, values_only=True))
    if len(surv) != 15 or len(survival_ci) != 15:
        raise ValueError(f"GPFM survival inventory changed: {len(surv)}/{len(survival_ci)}")
    for index, row in enumerate(surv):
        task, value = str(row[0]), row[9]
        _, low, high = survival_ci[index]
        add("Survival analysis", f"{task}/c-index", task, "c_index", value, f"95% CI [{low}, {high}]", "slide",
            "two-layer ABMIL survival head on frozen GPFM embeddings", "canonical_candidate", "public",
            "Public frozen-encoder survival leaf.", 7, "b")

    roi_ci = _gpfm_ci_groups(text, range(25, 37))
    wb8 = load_workbook(paths[8], read_only=True, data_only=True)
    roi = list(wb8["b"].iter_rows(min_row=2, values_only=True))
    if len(roi) != 16 or len(roi_ci) != 48:
        raise ValueError(f"GPFM ROI inventory changed: {len(roi)}/{len(roi_ci)}")
    for row_index, row in enumerate(roi):
        task = str(row[0])
        public = task in GPFM_PUBLIC_ROI
        for metric_index, (sheet, metric) in enumerate((("b", "balanced_accuracy"), ("c", "weighted_f1"), ("d", "auroc"))):
            value = wb8[sheet].cell(row_index + 2, 10).value
            _, low, high = roi_ci[row_index * 3 + metric_index]
            add("ROI classification", f"{task}/{metric}", task, metric, value, f"95% CI [{low}, {high}]", "tile",
                "linear classifier on frozen GPFM embeddings", "canonical_candidate" if public else "private_internal_excluded",
                "public" if public else "private_internal",
                "Public frozen linear-probe leaf." if public else "SAL/internal cohort is not publicly released.", 8, sheet)

    wb9 = load_workbook(paths[9], read_only=True, data_only=True)
    for row in wb9["a"].iter_rows(min_row=2, values_only=True):
        metric = {"ACC@1": "top1_accuracy", "ACC@3": "top3_accuracy", "ACC@5": "top5_accuracy"}[str(row[0])]
        add("ROI retrieval", f"CRC-100K/{metric}", "CRC-100K retrieval", metric, row[17], f"reported SD {row[18]}", "tile",
            "cosine nearest-neighbour retrieval on frozen GPFM embeddings", "canonical_candidate", "public",
            "Public frozen-embedding retrieval leaf.", 9, "a")
    for row in wb9["c"].iter_rows(min_row=2, values_only=True):
        metric = {"Open ACC": "open_accuracy", "Closed ACC": "closed_accuracy", "Overall ACC": "overall_accuracy"}[str(row[0])]
        add("PathVQA", f"PathVQA/{metric}", "PathVQA", metric, row[25], f"95% CI [{row[27]}, {row[26]}]", "tile",
            "task-specific VQA fine-tuning", "fine_tuned_excluded", "public",
            "Task-specific generative/VQA training is not a frozen-encoder probe.", 9, "c")
    for row in wb9["f"].iter_rows(min_row=2, values_only=True):
        task, metric = str(row[0]), _slug(str(row[1]))
        add("WSI report generation", f"{task}/{metric}", task, metric, row[18], f"reported SD {row[19]}", "slide",
            "task-specific report-generation fine-tuning", "fine_tuned_excluded", "public",
            "Task-specific generative training is not a frozen-encoder probe.", 9, "f")

    wb13 = load_workbook(paths[13], read_only=True, data_only=True)["a"]
    header13 = [str(v) if v is not None else "" for v in next(wb13.iter_rows(min_row=1, max_row=1, values_only=True))]
    gpfm13 = next(row for row in wb13.iter_rows(min_row=2, values_only=True) if str(row[0]).upper() == "GPFM")
    for col in (1, 2, 3, 4, 5, 6, 8):
        metric = _slug(header13[col])
        add("WSI-VQA", f"WSI-VQA/{metric}", "WSI-VQA", metric, gpfm13[col], "", "slide",
            "task-specific WSI-VQA fine-tuning", "fine_tuned_excluded", "public",
            "Task-specific generative/VQA training is not a frozen-encoder probe.", 13, "a")
    wb14 = load_workbook(paths[14], read_only=True, data_only=True)
    for sheet, organ in (("a", "lung"), ("b", "breast"), ("c", "kidney")):
        ws = wb14[sheet]
        header = [str(v) if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        row = next(row for row in ws.iter_rows(min_row=2, values_only=True) if str(row[0]).upper() == "GPFM")
        for col in range(1, 8):
            metric = _slug(header[col])
            add("Organ-specific WSI report generation", f"TCGA-{organ}/{metric}", f"TCGA-{organ}", metric, row[col], "", "slide",
                "task-specific report-generation fine-tuning", "fine_tuned_excluded", "public",
                "Task-specific generative training is not a frozen-encoder probe.", 14, sheet)

    for table, cells in ((1, (("balanced_accuracy", "0.752", "0.161"), ("weighted_f1", "0.736", "0.179"), ("auroc", "0.891", "0.096"))),
                         (19, (("c_index", "0.665", "0.071"),)),
                         (24, (("balanced_accuracy", "0.866", "0.136"), ("weighted_f1", "0.865", "0.142"), ("auroc", "0.946", "0.066")))):
        for metric, value, sd in cells:
            rows.append(_row(
                source_scope="official_primary_paper", source_table=f"Supplementary Table {table}",
                source_row=f"GPFM/average/{metric}", model_alias="GPFM", model_id="gpfm",
                model_revision=common["model_revision"], suite_id=common["suite_id"],
                task_label="cross-dataset average", evaluation_id=f"gpfm.nbe2025.t{table}.average.{metric}",
                dedup_key=f"gpfm.aggregate.t{table}.{metric}", metric=metric, value=value, value_unit="fraction",
                uncertainty=f"reported SD {sd}", level="mixed", magnification="mixed",
                embedding_recipe="aggregate across GPFM benchmark leaves", downstream_protocol="cross-dataset aggregate",
                cohort_access="mixed", reference_url=common["reference_url"],
                source_locator=f"Supplementary Information|Table {table}|GPFM",
                source_revision=common["source_revision"], source_sha256=GPFM_SUPP_SHA,
                inclusion_status="aggregate_excluded", inclusion_reason="Cross-dataset average is not an independent benchmark leaf.",
            ))
    if len(rows) != 224:
        raise ValueError(f"expected 224 GPFM observations, got {len(rows)}")
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--hibou-pdf", type=Path, required=True)
    parser.add_argument("--musk-supplement", type=Path, required=True)
    parser.add_argument("--gpfm-supplement", type=Path, required=True)
    parser.add_argument("--gpfm-workbook-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=Path("source_data"))
    args = parser.parse_args()
    _write(args.output_dir / "hibou_official_scores_2024.csv", hibou_rows(args.hibou_pdf))
    _write(args.output_dir / "musk_official_scores_2025.csv", musk_rows(args.musk_supplement))
    _write(args.output_dir / "gpfm_official_scores_2025.csv", gpfm_rows(args.gpfm_supplement, args.gpfm_workbook_dir))


if __name__ == "__main__":
    main()
